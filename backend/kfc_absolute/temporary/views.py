from datetime import datetime, timedelta

from django.http import HttpResponse
from django.shortcuts import redirect
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from rest_framework import generics, permissions

from users.models import User
from users.serializers import CustomUserAdminSerializer

from .mixins import UserFilterMixin
from .serializers import TemporaryCreateSerializer, TemporarySerializer
from .tasks import delete_record


class AdminView(generics.ListAPIView):
    """
    Temporary administrator preview
    """

    queryset = User.objects.all()
    serializer_class = CustomUserAdminSerializer
    permission_classes = (permissions.IsAdminUser,)


class ExportTemporaryDataView(generics.GenericAPIView):
    """
    Exporting data to Excel
    """

    queryset = User.objects.all()
    serializer_class = CustomUserAdminSerializer
    permission_classes = (permissions.IsAdminUser,)

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="temporary_data.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Temporary Data'

        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        headers = ['ФИО', 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье',
                   'Понедельник - Вторник', 'Вторник - Среда', 'Среда - Четверг', 'Четверг - Пятница',
                   'Пятница - Суббота', 'Суббота - Воскресенье', 'Воскресенье - Понедельник']

        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header_title)
            cell.alignment = Alignment(horizontal='center')

        row = 2
        for user_data in serializer.data:
            worksheet.cell(row=row, column=1, value=user_data['full_name'])
            temporary_data = user_data['temporary']
            for temp_data in temporary_data:
                day = temp_data.get('day', '')
                shift_type = temp_data.get('shift_type', '')
                col_index = headers.index(day) + 1
                worksheet.cell(row=row, column=col_index, value=shift_type)

                # Set "Смогу" if night is True
                if temp_data.get('night', False):
                    night_col_index = headers.index(day) + 8  # Get the corresponding "Ночные смены" column index
                    worksheet.cell(row=row, column=night_col_index, value='Смогу')

            row += 1

        # Auto-fit column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(response)
        return response


class TemporaryView(UserFilterMixin, generics.ListAPIView):
    """
    Viewing temporary
    """

    serializer_class = TemporarySerializer
    permission_classes = (permissions.IsAuthenticated,)


# class TemporaryDetailView(UserFilterMixin, generics.RetrieveUpdateDestroyAPIView):
#     """
#     Detailed temporal
#     """

#     serializer_class = TemporaryCreateSerializer
#     permission_classes = (permissions.IsAuthenticated,)


class TemporaryCreateView(UserFilterMixin, generics.CreateAPIView):
    """
    Creating temporary
    """

    serializer_class = TemporaryCreateSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        delete_record.apply_async(args=(serializer.instance.id,), eta=datetime.now() + timedelta(days=7))

        return redirect('/api/v1/temporary/create/')
